import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2.service_account import Credentials
import os
import tempfile
import traceback
from copy import copy
from .models import RegistroVehiculo

# Configuración de credenciales y archivo
CREDENTIALS_FILE = "C:/Users/ramiro_ravachino/Desktop/djangoproject/control_vehiculos/credentials/controlvehiculos-9ec3dd46ad04.json"
FILE_ID = "1BrF2FzLPlwHnCec1EFxZsDIi7MkOjaTHVZKU0xW6KtY"
SCOPES = ["https://www.googleapis.com/auth/drive"]

ENCABEZADOS = {
    "Vehiculos": [
        "ID", "DNI", "Entrada", "Apellido y Nombre", "Empresa", 
        "Ingreso Remitos", "Nro de Patente", "Egreso Remitos", "Salida"],
    "Camiones": [
        "ID", "Chofer", "DNI", "Tractor", "Semi", "Ingreso", 
        "Transporte", "Egreso", "Ingreso 2", "Egreso 2", "Ingreso 3", "Egreso 3"
    ],

    "Administrativos": [
        "ID", "Entrada", "Vehiculo", "Patente Nro", 
        "Apellido y Nombre", "Observaciones", "Salida"]
}

# Autenticación en Google Drive 
creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
drive_service = build("drive", "v3", credentials=creds)

# Mapeo entre los campos de la base de datos y los encabezados del Excel
MAPEO_CAMPOS = {
    "id": "ID",
    "dni": "DNI",   
    "entrada": "Entrada",
    "apellido_nombre": "Apellido y Nombre",
    "empresa": "Empresa",
    "ingreso_remitos": "Ingreso Remitos",
    "pte_numero": "Nro de Patente",
    "egreso_remitos": "Egreso Remitos",
    "salida_vehiculo": "Salida"
}
MAPEO_CAMPOS_ADMIN = {
    "id": "ID",
    "vehiculo": "Vehiculo",   
    "entrada_admin1": "Entrada",
    "apellido_nombre_admin": "Apellido y Nombre",
    "observaciones": "Observaciones",
    "entrada_admin2": "Entrada 2",
    "salida_admin1": "Salida",
    "salida_admin2": "Egreso 2",
    "entrada_admin2": "Entrada 2"
}
MAPEO_CAMPOS_CAMIONES = {
    "id": "ID",
    "chofer": "Chofer",   
    "dni": "DNI",
    "tractor": "Tractor",
    "semi": "Semi",
    "ingreso": "Ingreso",
    "transporte": "Transporte",
    "egreso": "Egreso",
    "ingreso2": "Ingreso 2",
    "ingreso3": "Ingreso 3",
    "egreso2": "Egreso 2",
    "egreso3": "Egreso 3",
}

def leer_datos_excel(sheet_name):
    """Lee los datos de una hoja específica del Excel y devuelve las filas como lista de diccionarios."""
    datos = []
    temp_file = None
    try:
        # Descargar archivo temporal
        temp_file = tempfile.mktemp(suffix=".xlsx")
        request = drive_service.files().export_media(
            fileId=FILE_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        with open(temp_file, "wb") as f:
            f.write(request.execute())

        # Cargar el archivo Excel
        workbook = openpyxl.load_workbook(temp_file)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Obtener encabezados exactamente como están en el Excel
            encabezados = [str(cell.value).strip() for cell in sheet[1] if cell.value]
            print(f"Encabezados encontrados en {sheet_name}: {encabezados}")  # Debug print

            # Leer datos
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # Si al menos una celda tiene datos
                    fila_dict = {}
                    for i, value in enumerate(row):
                        if i < len(encabezados):
                            fila_dict[encabezados[i]] = value
                    if any(fila_dict.values()):  # Solo agregar si hay al menos un valor
                        datos.append(fila_dict)
                        print(f"Fila leída de {sheet_name}: {fila_dict}")  # Debug print

        return datos
    except Exception as e:
        print(f"Error al leer el archivo Excel ({sheet_name}): {e}")
        traceback.print_exc()
        return []
    finally:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)


def limpiar_filas_vacias(sheet):
    """Limpia filas vacías al final de la hoja Excel."""
    max_row = sheet.max_row
    for row in range(max_row, 0, -1):  # Empezar desde la última fila
        if all(sheet.cell(row=row, column=col).value is None for col in range(1, sheet.max_column + 1)):
            sheet.delete_rows(row)  # Eliminar fila si está vacía
        else:
            break  # Detenernos al encontrar una fila con datos

def encontrar_primera_fila_vacia(sheet):
    """Devuelve la primera fila vacía real en una hoja."""
    for row in range(2, sheet.max_row + 1):  # Empezar desde la segunda fila (después de los encabezados)
        if all(sheet.cell(row=row, column=col).value is None for col in range(1, sheet.max_column + 1)):
            return row
    return sheet.max_row + 1  # Si no encuentra, devolver la siguiente fila


def copiar_estilo(origen, destino):
    """Copia el estilo de una celda a otra de forma segura."""
    if origen.has_style:
        if origen.font:
            destino.font = copy(origen.font)
        if origen.fill:
            destino.fill = copy(origen.fill)
        if origen.border:
            destino.border = copy(origen.border)
        if origen.alignment:
            destino.alignment = copy(origen.alignment)

from django.utils.timezone import localtime
import pytz
import traceback
import os
import tempfile
from openpyxl import load_workbook
from googleapiclient.http import MediaFileUpload
from datetime import datetime

def update_excel(data, sheet_name):
    temp_file = None
    try:
        # Descargar archivo temporal
        temp_file = tempfile.mktemp(suffix=".xlsx")
        request = drive_service.files().export_media(
            fileId=FILE_ID,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        with open(temp_file, "wb") as f:
            f.write(request.execute())

        workbook = load_workbook(temp_file)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        # Escribir encabezados si no existen
        if sheet.max_row == 0 or not all(sheet.cell(row=1, column=i+1).value for i in range(len(ENCABEZADOS[sheet_name]))):
            for col_num, header in enumerate(ENCABEZADOS[sheet_name], start=1):
                sheet.cell(row=1, column=col_num, value=header)

        # Buscar fila por ID o agregar una nueva
        id_col = ENCABEZADOS[sheet_name].index("ID") + 1
        found_row = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=id_col).value == data[0]:  # Comparar ID
                found_row = row
                break

        # Ajuste de zona horaria
        argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')

        # Convertir campos de fecha a la hora local antes de escribir
        for i in range(len(data)):
            if isinstance(data[i], datetime):
                data[i] = localtime(data[i], argentina_tz).strftime('%Y-%m-%d %H:%M')

        if found_row:
            for col_num, value in enumerate(data, start=1):
                if value:  # Solo sobrescribir si hay un valor
                    sheet.cell(row=found_row, column=col_num, value=value)
        else:
            # Usar la primera fila vacía para escribir datos nuevos
            new_row = encontrar_primera_fila_vacia(sheet)
            print(f"Agregando nueva fila: {new_row}")
            for col_num, value in enumerate(data, start=1):
                sheet.cell(row=new_row, column=col_num, value=value)

        # Guardar y cerrar el archivo antes de subirlo
        workbook.save(temp_file)
        workbook.close()  # Cerrar explícitamente el archivo

        # Subir el archivo actualizado a Google Drive
        media = MediaFileUpload(temp_file, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        drive_service.files().update(fileId=FILE_ID, media_body=media).execute()

    except Exception as e:
        print(f"Error al actualizar el archivo Excel: {e}")
        traceback.print_exc()

    finally:
        # Eliminar archivo temporal
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except PermissionError as pe:
                print(f"Error al eliminar el archivo temporal: {pe}")
