from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from googleapiclient.http import MediaFileUpload
import openpyxl
import tempfile

# Configurar las credenciales
CREDENTIALS_FILE = "C:/Users/ramiro_ravachino/Desktop/djangoproject/control_vehiculos/credentials/controlvehiculos-7769ce3e574f.json"
FILE_ID = "1BrF2FzLPlwHnCec1EFxZsDIi7MkOjaTHVZKU0xW6KtY"  # Reemplaza con el ID del archivo
creds = Credentials.from_service_account_file(
    CREDENTIALS_FILE, scopes=["https://www.googleapis.com/auth/drive"]
)
drive_service = build("drive", "v3", credentials=creds)

# Proceso de prueba: Descargar, modificar y subir
try:
    # 1. Descargar el archivo Excel
    temp_file = tempfile.mktemp(suffix=".xlsx")
    request = drive_service.files().export_media(
        fileId=FILE_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    with open(temp_file, "wb") as f:
        f.write(request.execute())
    print("Archivo descargado exitosamente.")

    # 2. Modificar el archivo
    workbook = openpyxl.load_workbook(temp_file)
    if "Prueba" not in workbook.sheetnames:
        workbook.create_sheet("Prueba")
    sheet = workbook["Prueba"]
    sheet.append(["Test Escritura", "Funciona", "Correcto"])
    workbook.save(temp_file)
    print("Datos agregados exitosamente.")

    # 3. Subir el archivo modificado
    media = MediaFileUpload(temp_file, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    drive_service.files().update(fileId=FILE_ID, media_body=media).execute()
    print("Archivo actualizado exitosamente en Google Drive.")

except Exception as e:
    print(f"Error durante la prueba: {e}")
